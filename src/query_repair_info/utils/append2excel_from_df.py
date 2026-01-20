import pandas as pd
from openpyxl import load_workbook
import os
import logging

logger = logging.getLogger(__name__)

def append_df_to_excel(df: pd.DataFrame, file_path: str, sheet_name: str = "Sheet1") -> None:
    if not os.path.exists(file_path):
        logger.error(f"The file {file_path} does not exist.")
        raise SystemExit(1)

    try:
        wb = load_workbook(file_path)

        if sheet_name in wb.sheetnames:
            startrow = wb[sheet_name].max_row
            header = False
        else:
            startrow = 0
            header = True
        
        with pd.ExcelWriter(file_path,engine="openpyxl",mode="a",if_sheet_exists="overlay") as writer:
                df.to_excel(writer,sheet_name=sheet_name,startrow=startrow,index=False,header=header)
        logger.info(f"Data appended to {file_path} in sheet {sheet_name} successfully.")
    except Exception as e:
        logger.error(f"Failed to append data to Excel file: {e}")
        raise
    finally:
        wb.close()