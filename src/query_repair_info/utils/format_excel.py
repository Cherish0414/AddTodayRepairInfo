from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, numbers, NamedStyle
import logging

logger = logging.getLogger(__name__)

def format_sheet(file_path: str, sheet_name: str) -> None:
    try:
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        
        style = NamedStyle(name="custom_style")
        style.font = Font(name='Yu Gothic')
        style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        width = {
            'A' : 18, # 修理前IMEI
            'B' : 18, # 修理後IMEI
            'C' : 14, # 管理番号
            'D' : 20, # 修理報告時間
            'E' : 14, # 修理パターン
            'F' : 14, # 修理方針
            'G' : 14, # 不具合の症状
            'H' : 14, # 責任
            'I' : 10, # TOK
            'J' : 10, # 在庫品交換
            'K' : 10, # 外装交換
            'L' : 10, # ミドルキャビ交換
            'M' : 10, # SIMトレー交換
            'N' : 10, # ドック交換
            'O' : 10, # フロントカバー
            'P' : 10, # 背面カバー
            'Q' : 10, # 基板交換
            'R' : 10, # SIMトレー・ドック交換
            'S' : 10, # その他
        }
        
        for col, col_width in width.items():
            ws.column_dimensions[col].width = col_width
            
        for row in ws.iter_rows(min_row=2, min_col=ws['A'][0].column, max_col=ws['B'][0].column):
            for cell in row:
                cell.number_format = numbers.FORMAT_NUMBER
        
        for row in ws.iter_rows(min_row=2, min_col=ws['C'][0].column, max_col=ws['C'][0].column):
            for cell in row:
                cell.number_format = numbers.FORMAT_TEXT
        
        for row in ws.iter_rows(min_row=2, min_col=ws['D'][0].column, max_col=ws['D'][0].column):
            for cell in row:
                cell.number_format = numbers.FORMAT_DATE_DATETIME
                
        for row in ws.iter_rows(min_row=2, min_col=ws['E'][0].column, max_col=ws['H'][0].column):
            for cell in row:
                cell.number_format = numbers.FORMAT_TEXT
        
        for row in ws.iter_rows(min_row=2, min_col=ws['I'][0].column, max_col=ws['S'][0].column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = style.font
                cell.border = style.border

        wb.save(file_path)
        logger.info(f"Formatted sheet '{sheet_name}' successfully.")

    except Exception as e:
        logger.error(f"Error formatting sheet '{sheet_name}': {e}")
        raise SystemExit(1)
    finally:
        wb.close()